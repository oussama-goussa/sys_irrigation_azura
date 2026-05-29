# ============================================================
# backend/services/user_service.py
# ============================================================

import uuid
import csv
import io
import openpyxl

from datetime import datetime, timezone
from sqlalchemy.orm import Session
from loguru import logger
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

from models.user_model import User, AuditLog
from core.security import hash_password, Role


# ── Audit log helper ──────────────────────────────────────────
def log_action(db: Session, username: str, action: str, detail: str = None, ip: str = None):
    entry = AuditLog(
        id       = str(uuid.uuid4()),
        username = username,
        action   = action,
        detail   = detail,
        ip       = ip,
    )
    db.add(entry)
    db.commit()


# ── User CRUD ─────────────────────────────────────────────────
def get_user(db: Session, username: str) -> User | None:
    return db.query(User).filter(User.username == username).first()


def get_all_users(db: Session) -> list:
    return [u.to_dict() for u in db.query(User).all()]


def create_user(
    db: Session, username: str, password: str,
    role: str, nom: str, email: str = None,
    farm_names=None
) -> User:
    user = User(
        username = username,
        password = hash_password(password),
        role     = role,
        nom      = nom,
        email    = email,
        farm_names = farm_names or [],
        actif    = True,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    logger.success(f"User créé : {username} ({role})")
    return user


def update_user(
    db: Session, username: str,
    nom: str = None, email: str = None, password: str = None,
    farm_names=None
) -> User | None:
    user = get_user(db, username)
    if not user:
        return None
    if nom is not None:
        user.nom = nom
    if email is not None:
        user.email = email
    if password is not None:
        user.password = hash_password(password)
    if farm_names is not None:
        user.farm_names = farm_names
    db.commit()
    db.refresh(user)
    logger.info(f"User mis à jour : {username}")
    return user


def update_user_role(db: Session, username: str, new_role: str) -> User | None:
    user = get_user(db, username)
    if not user:
        return None
    user.role = new_role
    db.commit()
    db.refresh(user)
    return user


def toggle_user_actif(db: Session, username: str) -> User | None:
    user = get_user(db, username)
    if not user:
        return None
    user.actif = not user.actif
    db.commit()
    db.refresh(user)
    return user


def update_last_login(db: Session, username: str):
    user = get_user(db, username)
    if user:
        user.last_login = datetime.now(timezone.utc)
        db.commit()


# ── Audit logs ────────────────────────────────────────────────
def get_audit_logs(db: Session, username: str = None, limit: int = 100) -> list:
    query = db.query(AuditLog).order_by(AuditLog.created_at.desc())
    if username:
        query = query.filter(AuditLog.username == username)
    return [log.to_dict() for log in query.limit(limit).all()]


# ── Export CSV ────────────────────────────────────────────────
def export_users_excel(db: Session) -> bytes:

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "azura_users"

    # Column widths — exact match to template
    for col, w in zip("ABCDEFG", [20.1, 30.0, 32.0, 16.9, 10.3, 33.8, 34.9]):
        ws.column_dimensions[col].width = w

    _s  = Side(border_style="thin", color="FF000000")
    _b  = Border(top=_s, bottom=_s, left=_s, right=_s)
    center = Alignment(horizontal="center", vertical="center")

    header_fill = PatternFill("solid", fgColor="FF275317")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    data_font   = Font(name="Calibri", size=11, bold=True, color="FF000000")

    # Header row
    ws.row_dimensions[1].height = 18.0
    for col_idx, label in enumerate(
        ["Nom d'utilisateur", "Nom complet", "Adresse e-mail", "Rôle", "Actif", "Dernière connexion", "Date de création"],
        start=1
    ):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = _b

    # Data rows
    for row_idx, u in enumerate(db.query(User).all(), start=2):
        ws.row_dimensions[row_idx].height = 15.0
        for col_idx, val in enumerate([
            u.username,
            u.nom,
            u.email or "",
            u.role,
            "oui" if u.actif else "non",
            str(u.last_login) if u.last_login else "",
            str(u.created_at),
        ], start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = data_font
            c.alignment = center
            c.border = _b

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ── Init default users ────────────────────────────────────────
import os

def init_default_users(db: Session):
    if db.query(User).count() > 0:
        logger.info("Users déjà initialisés")
        return

    admin_pwd = os.getenv("DEFAULT_ADMIN_PASSWORD")
    op_pwd    = os.getenv("DEFAULT_OPERATEUR_PASSWORD")

    if not admin_pwd or not op_pwd:
        raise RuntimeError(
            "DEFAULT_ADMIN_PASSWORD et DEFAULT_OPERATEUR_PASSWORD "
            "doivent être définis dans .env avant le premier démarrage"
        )

    for u in [
        {"username": "admin",     "password": admin_pwd, "role": Role.ADMIN,     "nom": "Administrateur Azura"},
        {"username": "operateur", "password": op_pwd,    "role": Role.OPERATEUR, "nom": "Opérateur Terrain Azura"},
    ]:
        create_user(db, u["username"], u["password"], u["role"], u["nom"])
    logger.success("Users par défaut créés ✅")
    # NE PAS logger les mots de passe