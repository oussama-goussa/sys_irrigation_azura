# ============================================================
# backend/routers/devices.py
# Endpoints lecture devices + sensor readings + export CSV
# Projet Azura Irrigation IA — GOUSSA Oussama
# ============================================================

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from datetime import datetime, date, timedelta
from typing import Optional

import csv, io
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from core.utils import filter_by_farm

from loguru import logger

from core.database import get_db
from core.security import require_any, require_operateur

from sqlalchemy import func, case
from datetime import datetime, timedelta

from models.sensor_model import (
    Device, SensorReading, IrrigationCycle,
    FertigationState, Alert, AlertThreshold
)

router = APIRouter(prefix="/api/devices", tags=["Devices & Capteurs"])

# Précharger les données en bulk avant la boucle
def _build_dashboard_bulk(devices: list, db: Session) -> list:
    if not devices:
        return []
    
    device_ids = [d.id for d in devices]
    since_24h  = datetime.utcnow() - timedelta(hours=24)
    
    # 1. Dernière lecture par device (une seule requête)
    subq = (
        db.query(
            SensorReading.device_id,
            func.max(SensorReading.timestamp).label("max_ts")
        )
        .filter(SensorReading.device_id.in_(device_ids))
        .group_by(SensorReading.device_id)
        .subquery()
    )
    last_readings = {
        sr.device_id: sr
        for sr in db.query(SensorReading)
        .join(subq, (SensorReading.device_id == subq.c.device_id) &
                    (SensorReading.timestamp  == subq.c.max_ts))
        .all()
    }
    
    # 2. Alertes actives par device (une seule requête)
    alert_counts = dict(
        db.query(Alert.device_id, func.count(Alert.id))
        .filter(Alert.device_id.in_(device_ids), Alert.resolved_at == None)
        .group_by(Alert.device_id)
        .all()
    )
    
    # 3. Lectures 24h par device (une seule requête)
    readings_24h = dict(
        db.query(SensorReading.device_id, func.count(SensorReading.id))
        .filter(SensorReading.device_id.in_(device_ids),
                SensorReading.timestamp >= since_24h)
        .group_by(SensorReading.device_id)
        .all()
    )
    
    # 4. Seuils par device (une seule requête)
    all_thresholds: dict[int, dict] = {}
    for t in db.query(AlertThreshold).filter(
        AlertThreshold.device_id.in_(device_ids),
        AlertThreshold.is_active == True
    ).all():
        all_thresholds.setdefault(t.device_id, {})[t.parameter] = t
    
    # Construire les summaries sans requêtes supplémentaires
    result = []
    for d in devices:
        last = last_readings.get(d.id)
        thresholds = all_thresholds.get(d.id, {})
        
        online = False
        last_seen_min = None
        if last and last.timestamp:
            delta = (datetime.utcnow() - last.timestamp).total_seconds() / 60
            last_seen_min = int(delta)
            online = delta < 15
        
        def get_thresh(param):
            t = thresholds.get(param)
            return (t.threshold_min if t else None, t.threshold_max if t else None)
        
        metrics = {}
        if last:
            for key, attr, p in [
                ("ec",   "ec_actual",  "ec_actual"),
                ("ph",   "ph_actual",  "ph_actual"),
                ("temp", "avg_temp",   "avg_temp"),
                ("hum",  "humidity",   "humidity"),
            ]:
                val = getattr(last, attr, None)
                mn, mx = get_thresh(p)
                metrics[key] = {"value": val, "min": mn, "max": mx,
                                 "status": _status_color(val, mn, mx)}
            metrics["rad"]  = {"value": last.radiation,  "status": "ok"}
            metrics["flow"] = {"value": last.flow,        "status": "ok"}
        
        result.append({
            "id": d.id, "farm_name": d.farm_name, "house_number": d.house_number,
            "room_number": d.room_number, "device_id": d.device_id,
            "is_active": d.is_active, "online": online,
            "last_seen_min": last_seen_min,
            "last_timestamp": str(last.timestamp) if last else None,
            "metrics": metrics,
            "irrigation_active": False,
            "active_alerts": alert_counts.get(d.id, 0),
            "readings_24h": readings_24h.get(d.id, 0),
        })
    
    return result

def _check_device_access(device: Device, user: dict):
    """Lève 403 si l'utilisateur n'a pas accès à ce device."""
    if user["role"] == "admin":
        return
    allowed = user.get("farm_names", [])
    if device.farm_name not in allowed:
        raise HTTPException(
            status_code=403,
            detail="Accès refusé à ce device"
        )
    
# ── helpers ───────────────────────────────────────────────────

def _status_color(value, tmin, tmax):
    """Retourne 'ok' | 'warning' | 'critical'"""
    if value is None:
        return "unknown"
    if tmin is not None and value < tmin:
        return "critical"
    if tmax is not None and value > tmax:
        # Proche du max (90%) → warning
        if tmax > 0 and value > tmax * 0.90:
            return "warning"
        return "critical"
    return "ok"


def _build_device_summary(device: Device, db: Session) -> dict:
    """Construit le résumé complet d'un device pour le dashboard."""

    # Dernière lecture
    last = (
        db.query(SensorReading)
        .filter(SensorReading.device_id == device.id)
        .order_by(desc(SensorReading.timestamp))
        .first()
    )

    # Dernière lecture il y a combien de minutes ?
    online = False
    last_seen_min = None
    if last and last.timestamp:
        delta = datetime.utcnow() - last.timestamp
        last_seen_min = int(delta.total_seconds() / 60)
        online = last_seen_min < 15  # hors-ligne si pas de données depuis 15 min

    # Récupérer les seuils configurés pour ce device
    thresholds = {
        t.parameter: t
        for t in db.query(AlertThreshold)
        .filter(AlertThreshold.device_id == device.id, AlertThreshold.is_active == True)
        .all()
    }

    def get_thresh(param):
        t = thresholds.get(param)
        return (t.threshold_min if t else None, t.threshold_max if t else None)

    # Dernière lecture programme irrigation
    cycle = (
        db.query(IrrigationCycle)
        .filter(IrrigationCycle.device_id == device.id)
        .order_by(desc(IrrigationCycle.timestamp))
        .first()
    )

    # Alertes actives non résolues
    active_alerts = (
        db.query(Alert)
        .filter(Alert.device_id == device.id, Alert.resolved_at == None)
        .count()
    )

    # Nombre de lectures dans les 24h
    since_24h = datetime.utcnow() - timedelta(hours=24)
    readings_24h = (
        db.query(func.count(SensorReading.id))
        .filter(SensorReading.device_id == device.id,
                SensorReading.timestamp >= since_24h)
        .scalar()
    )

    ec_min, ec_max   = get_thresh("ec_actual")
    ph_min, ph_max   = get_thresh("ph_actual")
    tmp_min, tmp_max = get_thresh("avg_temp")
    hum_min, hum_max = get_thresh("humidity")

    metrics = {}
    if last:
        metrics = {
            "ec":    {"value": last.ec_actual,   "min": ec_min,  "max": ec_max,  "status": _status_color(last.ec_actual,   ec_min,  ec_max)},
            "ph":    {"value": last.ph_actual,   "min": ph_min,  "max": ph_max,  "status": _status_color(last.ph_actual,   ph_min,  ph_max)},
            "temp":  {"value": last.avg_temp,    "min": tmp_min, "max": tmp_max, "status": _status_color(last.avg_temp,    tmp_min, tmp_max)},
            "hum":   {"value": last.humidity,    "min": hum_min, "max": hum_max, "status": _status_color(last.humidity,    hum_min, hum_max)},
            "rad":   {"value": last.radiation,   "min": None,    "max": None,    "status": "ok"},
            "flow":  {"value": last.flow,        "min": None,    "max": None,    "status": "ok"},
            "wind":  {"value": last.wind_speed,  "min": None,    "max": None,    "status": "ok"},
            "vpd":   {"value": last.vpd,         "min": None,    "max": None,    "status": "ok"},
        }

    # Statut irrigation actif ?
    irrigation_active = False
    if cycle:
        irrigation_active = str(cycle.irrigation_active or "0") not in ("0", "false", "False", "")

    return {
        "id"                : device.id,
        "farm_name"         : device.farm_name,
        "house_number"      : device.house_number,
        "room_number"       : device.room_number,
        "controller_type"   : device.controller_type,
        "controller_version": device.controller_version,
        "device_id"         : device.device_id,
        "is_active"         : device.is_active,
        "online"            : online,
        "last_seen_min"     : last_seen_min,
        "last_timestamp"    : str(last.timestamp) if last else None,
        "metrics"           : metrics,
        "irrigation_active" : irrigation_active,
        "active_alerts"     : active_alerts,
        "readings_24h"      : readings_24h,
    }


# Ajouter après les imports existants, dans le router
@router.get("/farms")
def list_farms(
    db: Session = Depends(get_db),
    user = Depends(require_any)
):
    """Retourne la liste des fermes distinctes (pour le select lors de création user)."""
    farms = db.query(Device.farm_name).filter(Device.is_active == True).distinct().order_by(Device.farm_name).all()
    return [f[0] for f in farms]

# ── GET /api/devices — Liste toutes les fermes groupées ──────
@router.get("")
def list_devices(
    db  : Session = Depends(get_db),
    user        = Depends(require_any)
):
    """
    Retourne la liste de tous les devices groupés par farm_name.
    Utilisé par la sidebar et le dashboard global.
    """
    devices = db.query(Device).filter(Device.is_active == True).all()

    # Après la query devices, ajouter :
    if user["role"] != "admin":
        allowed = user.get("farm_names", [])
        devices = [d for d in devices if d.farm_name in allowed]

    # Grouper par ferme
    farms = {}
    for d in devices:
        farm = d.farm_name
        if farm not in farms:
            farms[farm] = []
        farms[farm].append({
            "id"          : d.id,
            "farm_name"   : d.farm_name,
            "house_number": d.house_number,
            "room_number" : d.room_number,
            "device_id"   : d.device_id,
            "controller_type"   : d.controller_type,    # ← AJOUTER
            "controller_version": d.controller_version, # ← AJOUTER
            "is_active"   : d.is_active,
        })

    result = []
    for farm_name, houses in sorted(farms.items()):
        result.append({
            "farm_name": farm_name,
            "houses"   : sorted(houses, key=lambda x: x["house_number"]),
        })

    return result


# ── GET /api/devices/dashboard — Résumé toutes serres ────────
@router.get("/dashboard")
def dashboard_summary(
    db  : Session = Depends(get_db),
    user        = Depends(require_any)
):
    """
    Dashboard global : résumé de toutes les fermes/serres.
    Retourne les métriques temps réel, statut online/offline, alertes.
    """
    devices = db.query(Device).filter(Device.is_active == True).all()

    # Après la query devices, ajouter :
    if user["role"] != "admin":
        allowed = user.get("farm_names", [])
        devices = [d for d in devices if d.farm_name in allowed]

    farms = {}
    total_online  = 0
    total_alerts  = 0
    total_readings_24h = 0

    summaries = _build_dashboard_bulk(devices, db)

    for summary in summaries:
        if summary["online"]:     total_online += 1
        total_alerts              += summary["active_alerts"]
        total_readings_24h        += summary["readings_24h"]
        farm = summary["farm_name"]
        if farm not in farms:
            farms[farm] = {"farm_name": farm, "houses": []}
        farms[farm]["houses"].append(summary)

    farms_list = sorted(
        [{"farm_name": k, "houses": sorted(v["houses"], key=lambda x: x["house_number"])}
         for k, v in farms.items()],
        key=lambda x: x["farm_name"]
    )

    return {
        "farms"            : farms_list,
        "stats"            : {
            "total_farms"  : len(farms),
            "total_houses" : len(devices),
            "online_count" : total_online,
            "offline_count": len(devices) - total_online,
            "active_alerts": total_alerts,
            "readings_24h" : total_readings_24h,
        },
        "generated_at": datetime.utcnow().isoformat(),
    }


# ── GET /api/devices/{id}/latest — Dernière lecture ──────────
@router.get("/{device_id}/latest")
def get_latest(
    device_id: int,
    db       : Session = Depends(get_db),
    user             = Depends(require_any)
):
    """
    Retourne la dernière lecture capteurs + état pompes/vannes.
    Utilisé pour les StatCards (rafraîchissement 30s).
    """
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="Device non trouvé")
    _check_device_access(device, user)

    last_sr = (
        db.query(SensorReading)
        .filter(SensorReading.device_id == device_id)
        .order_by(desc(SensorReading.timestamp))
        .first()
    )

    last_cycle = (
        db.query(IrrigationCycle)
        .filter(IrrigationCycle.device_id == device_id)
        .order_by(desc(IrrigationCycle.timestamp))
        .first()
    )

    last_fert = (
        db.query(FertigationState)
        .filter(FertigationState.device_id == device_id)
        .order_by(desc(FertigationState.timestamp))
        .first()
    )

    # Lecture 30 min avant pour calculer les deltas
    thirty_min_ago = datetime.utcnow() - timedelta(minutes=30)
    prev_sr = (
        db.query(SensorReading)
        .filter(SensorReading.device_id == device_id,
                SensorReading.timestamp <= thirty_min_ago)
        .order_by(desc(SensorReading.timestamp))
        .first()
    )

    def delta(curr, prev, attr):
        c = getattr(curr, attr, None) if curr else None
        p = getattr(prev, attr, None) if prev else None
        if c is None or p is None:
            return None
        return round(c - p, 3)

    thresholds = {
        t.parameter: {"min": t.threshold_min, "max": t.threshold_max}
        for t in db.query(AlertThreshold)
        .filter(AlertThreshold.device_id == device_id, AlertThreshold.is_active == True)
        .all()
    }

    def sr_dict(sr):
        if not sr:
            return None
        return {
            "timestamp"     : str(sr.timestamp),
            "ec_actual"     : sr.ec_actual,
            "ph_actual"     : sr.ph_actual,
            "avg_temp"      : sr.avg_temp,
            "humidity"      : sr.humidity,
            "radiation"     : sr.radiation,
            "radiation_sum" : sr.radiation_sum,
            "flow"          : sr.flow,
            "flow_nominal"  : sr.flow_nominal,
            "wind_speed"    : sr.wind_speed,
            "wind_dir"      : sr.wind_dir,
            "vpd"           : sr.vpd,
            "vpd_sum"       : sr.vpd_sum,
            "daily_rain"    : sr.daily_rain,
            "ec_ph_status"  : sr.ec_ph_status,
            "outside_temp"  : sr.outside_temp,
            "outside_humidity": sr.outside_humidity,
        }

    def cycle_dict(c):
        if not c:
            return None
        return {
            "timestamp"        : str(c.timestamp),
            "sequence"         : c.sequence,
            "cycle_prog"       : c.cycle_prog,
            "cycle_act"        : c.cycle_act,
            "next_sequence"    : c.next_sequence,
            "next_seq_time"    : c.next_seq_time,
            "remaining_time"   : c.remaining_time,
            "active_order"     : c.active_order,
            "pump1": c.pump1, "pump2": c.pump2, "pump3": c.pump3,
            "pump4": c.pump4, "pump5": c.pump5, "pump6": c.pump6,
            "main_valve1": c.main_valve1, "main_valve2": c.main_valve2,
            "main_valve3": c.main_valve3, "main_valve4": c.main_valve4,
            "main_valve5": c.main_valve5, "main_valve6": c.main_valve6,
            "valve1": c.valve1, "valve2": c.valve2,
            "valve3": c.valve3, "valve4": c.valve4,
            "valves_in_irrig"  : c.valves_in_irrig,
            "irrigation_active": c.irrigation_active,
            "fert_active"      : c.fert_active,
            "booster_active"   : c.booster_active,
            "misting_active"   : c.misting_active,
            "cooling_active"   : c.cooling_active,
            "flushing_active"  : c.flushing_active,
            "water_mode"       : c.water_mode,
            "water_prg_qty"    : c.water_prg_qty,
            "water_prg_time"   : c.water_prg_time,
            "water_act_qty"    : c.water_act_qty,
            "water_act_time"   : c.water_act_time,
            "water_left"       : c.water_left,
            "fertilizer_qty"   : c.fertilizer_qty,
            "dosing_pump_type1": c.dosing_pump_type1,
            "dosing_pump_type2": c.dosing_pump_type2,
            "dosing_pump_type3": c.dosing_pump_type3,
            "dosing_pump_type4": c.dosing_pump_type4,
            "dosing_pump_type5": c.dosing_pump_type5,
            "dosing_pump_type6": c.dosing_pump_type6,
            "dosing_pump_type7": c.dosing_pump_type7,
            "dosing_pump_type8": c.dosing_pump_type8,
            "valve_prog": c.valve_prog, "fert_prog": c.fert_prog,
            "manual_prog": c.manual_prog, "pause": c.pause,
        }

    def fert_dict(f):
        if not f:
            return None
        d = {"timestamp": str(f.timestamp)}
        for i in range(1, 9):
            d[f"fert_open{i}"] = getattr(f, f"fert_open{i}")
            d[f"fert_min{i}"]  = getattr(f, f"fert_min{i}")
            d[f"fert_act{i}"]  = getattr(f, f"fert_act{i}")
            d[f"fert_max{i}"]  = getattr(f, f"fert_max{i}")
            d[f"fert_flow{i}"] = getattr(f, f"fert_flow{i}")
        return d

    delta_min = None   # ← initialiser avant le if
    online = False
    if last_sr and last_sr.timestamp:
        ts = last_sr.timestamp.replace(tzinfo=None) if last_sr.timestamp.tzinfo else last_sr.timestamp
        delta_min = (datetime.utcnow() - ts).total_seconds() / 60
        online = delta_min < 15

    return {
        "device"    : device.to_dict(),
        "online"    : online,
        "sensor"    : sr_dict(last_sr),
        "deltas"    : {
            "ec"   : delta(last_sr, prev_sr, "ec_actual"),
            "ph"   : delta(last_sr, prev_sr, "ph_actual"),
            "temp" : delta(last_sr, prev_sr, "avg_temp"),
            "hum"  : delta(last_sr, prev_sr, "humidity"),
        },
        "thresholds": thresholds,
        "last_seen_min": int(delta_min) if delta_min is not None else None,
        "cycle"     : cycle_dict(last_cycle),
        "fertigation": fert_dict(last_fert),
    }


# ── GET /api/devices/{id}/history — Historique paginé ────────
@router.get("/{device_id}/history")
def get_history(
    device_id : int,
    date_from : Optional[str] = Query(None, regex=r'^\d{4}-\d{2}-\d{2}$'),
    date_to   : Optional[str] = Query(None, regex=r'^\d{4}-\d{2}-\d{2}$'),
    page      : int = Query(1, ge=1, le=1000),
    per_page  : int = Query(50, ge=10, le=500),   # max 500 au lieu de 5000
    db        : Session = Depends(get_db),
    user               = Depends(require_any)
):
    """
    Historique paginé des lectures capteurs pour le tableau.
    """
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="Device non trouvé")
    _check_device_access(device, user)

    # Période par défaut : aujourd'hui
    if not date_from:
        date_from = date.today().isoformat()
    if not date_to:
        date_to = date.today().isoformat()

    try:
        dt_from = datetime.strptime(date_from, "%Y-%m-%d")
        dt_to   = datetime.strptime(date_to,   "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    except ValueError:
        raise HTTPException(status_code=400, detail="Format date invalide (YYYY-MM-DD)")
    
    # Limiter la plage à 90 jours
    if (dt_to - dt_from).days > 90:
        raise HTTPException(status_code=400, detail="Plage limitée à 90 jours")

    q = (
        db.query(SensorReading)
        .filter(
            SensorReading.device_id == device_id,
            SensorReading.timestamp >= dt_from,
            SensorReading.timestamp <= dt_to,
        )
        .order_by(desc(SensorReading.timestamp))
    )

    total = q.count()
    rows  = q.offset((page - 1) * per_page).limit(per_page).all()

    thresholds = {
        t.parameter: {"min": t.threshold_min, "max": t.threshold_max}
        for t in db.query(AlertThreshold)
        .filter(AlertThreshold.device_id == device_id, AlertThreshold.is_active == True)
        .all()
    }

    data = []
    for r in rows:
        data.append({
            "timestamp"     : str(r.timestamp),
            "ec_actual"     : r.ec_actual,
            "ph_actual"     : r.ph_actual,
            "avg_temp"      : r.avg_temp,
            "humidity"      : r.humidity,
            "radiation"     : r.radiation,
            "radiation_sum" : r.radiation_sum,
            "flow"          : r.flow,
            "flow_nominal"  : r.flow_nominal,
            "ec_ph_status"  : r.ec_ph_status,
            "vpd"           : r.vpd,
            "wind_speed"    : r.wind_speed,
            "daily_rain"    : r.daily_rain,
            "outside_temp"     : r.outside_temp,
            "outside_humidity" : r.outside_humidity,
        })

    return {
        "device_id" : device_id,
        "date_from" : date_from,
        "date_to"   : date_to,
        "total"     : total,
        "page"      : page,
        "per_page"  : per_page,
        "pages"     : max(1, (total + per_page - 1) // per_page),
        "thresholds": thresholds,
        "data"      : data,
    }


# ── GET /api/devices/{id}/export — Export CSV ─────────────────
@router.get("/{device_id}/export")
def export_history_excel(
    device_id : int,
    date_from : Optional[str] = Query(None),
    date_to   : Optional[str] = Query(None),
    db        : Session = Depends(get_db),
    user               = Depends(require_operateur)
):
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="Device non trouvé")
    _check_device_access(device, user)

    if not date_from:
        date_from = date.today().isoformat()
    if not date_to:
        date_to = date.today().isoformat()

    dt_from = datetime.strptime(date_from, "%Y-%m-%d")
    dt_to   = datetime.strptime(date_to,   "%Y-%m-%d").replace(hour=23, minute=59, second=59)

    rows = (
        db.query(SensorReading)
        .filter(
            SensorReading.device_id == device_id,
            SensorReading.timestamp >= dt_from,
            SensorReading.timestamp <= dt_to,
        )
        .order_by(desc(SensorReading.timestamp))
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suivi Irrigation"

    # Column widths — exact match to template
    for col, w in zip("ABCDEFGHI", [20.0, 12, 9.66, 9.0, 10.66, 16.0, 13.0, 10.77, 12.88]):
        ws.column_dimensions[col].width = w

    header_fill  = PatternFill("solid", fgColor="FF404040")
    header_font  = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    data_font    = Font(name="Calibri", size=11, bold=True)
    center       = Alignment(horizontal="center", vertical="center")
    _s   = Side(border_style="thin", color="FF000000")
    _b   = Border(top=_s, bottom=_s, left=_s, right=_s)

    # Header row
    ws.row_dimensions[1].height = 18.0
    for col_idx, label in enumerate(
        ["Timestamp", "Ferme", "Station", "EC", "pH", "Température", "Humidité", "Rad", "Débit"],
        start=1
    ):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = _b

    # Data rows
    for row_idx, r in enumerate(rows, start=2):
        ws.row_dimensions[row_idx].height = 15.0
        for col_idx, val in enumerate([
            r.timestamp,
            device.farm_name,
            device.house_number,
            r.ec_actual,
            r.ph_actual,
            r.avg_temp,
            r.humidity,
            r.radiation,
            r.flow,
        ], start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = data_font
            c.alignment = center
            c.border = _b
            if col_idx == 1 and val:
                c.number_format = "DD/MM/YYYY HH:MM:SS"

    output = BytesIO()
    wb.save(output)
    excel_bytes = output.getvalue()

    filename = f"{device.farm_name}_Station{device.house_number}_{date_from}_{date_to}.xlsx"
    from fastapi.responses import Response
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(excel_bytes)),
        },
    )

# ── GET /api/devices/{id}/alerts — Alertes actives ───────────
@router.get("/{device_id}/alerts")
def get_alerts(
    device_id: int,
    resolved : bool = Query(False),
    limit    : int  = Query(50),
    db       : Session = Depends(get_db),
    user             = Depends(require_any)
):
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="Device non trouvé")
    _check_device_access(device, user)    
    
    q = db.query(Alert).filter(Alert.device_id == device_id)
    if not resolved:
        q = q.filter(Alert.resolved_at == None)
    alerts = q.order_by(desc(Alert.timestamp)).limit(limit).all()

    # Charger les seuils configurés pour enrichir les alertes sans min/max
    thresholds = {
        t.parameter.upper(): t
        for t in db.query(AlertThreshold)
        .filter(AlertThreshold.device_id == device_id, AlertThreshold.is_active == True)
        .all()
    }

    result = []
    for a in alerts:
        d = a.to_dict()
        # Si l'alerte n'a pas de seuils enregistrés → les prendre depuis AlertThreshold
        if d.get("threshold_min") is None and d.get("threshold_max") is None:
            t = thresholds.get((a.alert_type or "").upper())
            if t:
                d["threshold_min"] = t.threshold_min
                d["threshold_max"] = t.threshold_max
        result.append(d)

    return result

# ── GET /api/devices/{id}/thresholds — Seuils configurés ─────
@router.get("/{device_id}/thresholds")
def get_thresholds(
    device_id: int,
    db       : Session = Depends(get_db),
    user             = Depends(require_any)
):
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="Device non trouvé")
    _check_device_access(device, user)

    """Retourne les seuils d'alerte configurés pour un device."""
    thresholds = (
        db.query(AlertThreshold)
        .filter(AlertThreshold.device_id == device_id)
        .all()
    )
    return [
        {
            "id"           : t.id,
            "parameter"    : t.parameter,
            "threshold_min": t.threshold_min,
            "threshold_max": t.threshold_max,
            "severity"     : t.severity,
            "is_active"    : t.is_active,
        }
        for t in thresholds
    ]

# ── GET /api/devices/{id}/tours — Tours d'irrigation ─────────
@router.get("/{device_id}/tours")
def get_tours(
    device_id : int,
    date      : Optional[str] = Query(None, description="YYYY-MM-DD"),
    db        : Session = Depends(get_db),
    user               = Depends(require_any)
):
    """
    Retourne les tours d'irrigation d'un device pour une date.
    Par défaut : aujourd'hui.
    """
    from models.sensor_model import IrrigationTour

    target_date = date if date else datetime.utcnow().date().isoformat()

    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="Device non trouvé")
    _check_device_access(device, user)

    tours = (
        db.query(IrrigationTour)
        .filter(
            IrrigationTour.device_id == device_id,
            IrrigationTour.date      == target_date,
        )
        .order_by(IrrigationTour.tour_num.asc())
        .all()
    )

    # Toujours retourner 16 slots
    slots = []
    tours_dict = {t.tour_num: t for t in tours}

    for i in range(1, 30):
        t = tours_dict.get(i)
        if t:
            slots.append({
                "tour_num"       : i,
                "debut"          : t.debut.strftime('%H:%M') if t.debut else None,
                "fin"            : t.fin.strftime('%H:%M') if t.fin else None,
                "duree_min"      : t.duree_min,
                "prg_time_min"   : t.prg_time_min,
                "repos_apres_min": t.repos_apres_min,
                "is_complete"    : t.is_complete,
                "v_apport"       : t.v_apport,
                "ec_apport"      : t.ec_apport,
                "ph_apport"      : t.ph_apport,
                "radiation_sum"  : t.radiation_sum,
                "cumul_radiation": t.cumul_radiation, 
            })
        else:
            slots.append({
                "tour_num"       : i,
                "debut"          : None,
                "fin"            : None,
                "duree_min"      : None,
                "prg_time_min"   : None,
                "repos_apres_min": None,
                "is_complete"    : False,
                "v_apport"       : None,
                "ec_apport"      : None,
                "ph_apport"      : None,
                "radiation_sum"  : None,
                "cumul_radiation": None,
            })

    return {
        "device_id"  : device_id,
        "farm_name"  : device.farm_name,
        "house_number": device.house_number,
        "date"       : target_date,
        "tours"      : slots,
        "total_tours": len(tours),
    }

@router.patch("/{device_id}/alerts/{alert_id}/resolve")
def resolve_alert(
    device_id: int,
    alert_id: int,
    db: Session = Depends(get_db),
    user = Depends(require_any)
):
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="Device non trouvé")
    _check_device_access(device, user)

    from datetime import datetime
    alert = db.query(Alert).filter(
        Alert.id == alert_id,
        Alert.device_id == device_id
    ).first()
    if not alert:
        raise HTTPException(404, "Alerte non trouvée")
    alert.resolved_at = datetime.utcnow()
    alert.resolved_by = user["username"]
    db.commit()
    return {"message": "Alerte résolue ✅"}

# ── POST /api/devices/check-offline — Vérifier stations hors ligne ──
@router.post("/check-offline")
def check_offline_stations(
    db   : Session = Depends(get_db),
    user           = Depends(require_any),
):
    """
    Vérifie toutes les stations et crée des alertes OFFLINE
    si aucune donnée depuis plus de 20 minutes.
    Appelé automatiquement par le frontend toutes les 5 min.
    """
    from core.security import _redis   # réutiliser le client Redis existant
    
    # Verrou distribué : un seul check à la fois (expire en 60s)
    lock_key = "lock:check_offline"
    if _redis and not _redis.set(lock_key, "1", nx=True, ex=60):
        return {"status": "skipped", "reason": "already_running"}
    
    try:
        from datetime import timedelta

        cutoff_offline = datetime.utcnow() - timedelta(minutes=20)
        cutoff_alert   = datetime.utcnow() - timedelta(minutes=25)

        devices = db.query(Device).filter(Device.is_active == True).all()

        if user["role"] != "admin":
            allowed = user.get("farm_names", [])
            devices = [d for d in devices if d.farm_name in allowed]

        created = 0
        resolved = 0

        for device in devices:
            last = (
                db.query(SensorReading)
                .filter(SensorReading.device_id == device.id)
                .order_by(desc(SensorReading.timestamp))
                .first()
            )

            is_offline = (last is None) or (last.timestamp < cutoff_offline)

            if is_offline:
                # Pas d'alerte OFFLINE non résolue récente ?
                existing = db.query(Alert).filter(
                    Alert.device_id  == device.id,
                    Alert.alert_type == "OFFLINE",
                    Alert.resolved_at == None,
                    Alert.timestamp  >= cutoff_alert,
                ).first()

                if existing:
                    if last:
                        minutes_ago = int((datetime.utcnow() - last.timestamp).total_seconds() / 60)
                        d, h, m = minutes_ago // 1440, (minutes_ago % 1440) // 60, minutes_ago % 60
                        if d > 0:
                            duration_str = f"{d}j {h}h" if h else f"{d}j"
                        elif h > 0:
                            duration_str = f"{h}h {m}min" if m else f"{h}h"
                        else:
                            duration_str = f"{minutes_ago} min"
                        existing.message = (
                            f"Station hors ligne depuis {duration_str} — "
                            f"{device.farm_name} Station {device.house_number}"
                        )
                        existing.value_detected = float(minutes_ago)
                else:
                    if last:
                        minutes_ago = int((datetime.utcnow() - last.timestamp).total_seconds() / 60)
                        d, h, m = minutes_ago // 1440, (minutes_ago % 1440) // 60, minutes_ago % 60
                        if d > 0:
                            duration_str = f"{d}j {h}h" if h else f"{d}j"
                        elif h > 0:
                            duration_str = f"{h}h {m}min" if m else f"{h}h"
                        else:
                            duration_str = f"{minutes_ago} min"
                        msg = (
                            f"Station hors ligne depuis {duration_str} — "
                            f"{device.farm_name} Station {device.house_number}"
                        )
                    else:
                        minutes_ago = None
                        msg = f"Station jamais connectée — {device.farm_name} Station {device.house_number}"

                    alert = Alert(
                        device_id      = device.id,
                        timestamp      = datetime.utcnow(),
                        alert_type     = "OFFLINE",
                        value_detected = float(minutes_ago) if minutes_ago else None,
                        threshold_min  = None,
                        threshold_max  = 20.0,
                        severity       = "CRITICAL",
                        message        = msg,
                    )
                    db.add(alert)
                    created += 1
                    logger.warning(f"⚠️ OFFLINE : {msg}")
            else:
                # Station revenue en ligne → résoudre alertes OFFLINE existantes
                nb = db.query(Alert).filter(
                    Alert.device_id  == device.id,
                    Alert.alert_type == "OFFLINE",
                    Alert.resolved_at == None,
                ).update({
                    "resolved_at": datetime.utcnow(),
                    "resolved_by": "auto-recover",
                })
                resolved += nb

        db.commit()
        return {
            "checked" : len(devices),
            "offline_alerts_created": created,
            "auto_resolved": resolved,
        }
        pass
    finally:
        if _redis:
            _redis.delete(lock_key)