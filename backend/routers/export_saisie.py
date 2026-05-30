# ============================================================
# backend/routers/export_saisie.py
# Export Excel — Suivi Irrigation (structure exacte du template)
# ============================================================

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from io import BytesIO
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.saisie_model import SaisieJournaliere

from core.database import get_db
from core.security import get_current_user

router = APIRouter(prefix="/api/export", tags=["Export"])

# ── Colors (exact from template) ─────────────────────────────
GREEN_LIGHT  = "FF99FF66"   # Radiation / Cumul Radiation
GREEN_MED    = "FF70AD47"   # Heure, Durée, V Apport, EC, pH, V Drain, EC/pH Drain
ORANGE       = "FFF4B942"   # % Drain, Moy % Drain, Heure Matin/Soir, Poids, Ressuyage, EC Bassin, Summary
WHITE        = "FFFFFFFF"
DARK_HEADER  = "FF404040"   # Header row background

# ── Helpers ───────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=True, size=10, color="FF000000"):
    return Font(bold=bold, size=size, name="Calibri", color=color)

def _align():
    return Alignment(horizontal="center", vertical="center")

def _border_thin():
    s = Side(border_style="thin", color="FF000000")
    return Border(top=s, bottom=s, left=s, right=s)

def _border_label():
    """Label col F: thin on top, bottom, right"""
    s = Side(border_style="thin", color="FF000000")
    return Border(top=s, bottom=s, right=s)

def _set(ws, row, col, value, fill_hex=None, bold=True, size=10,
         font_color="FF000000", border=None, num_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, size=size, color=font_color)
    cell.alignment = _align()
    if fill_hex:
        cell.fill = _fill(fill_hex)
    if border:
        cell.border = border
    if num_format:
        cell.number_format = num_format
    return cell

def _parse_time_str(s):
    """'HH:MM' or 'HH:MM:SS' → datetime.time or None"""
    if not s:
        return None
    try:
        parts = str(s).split(":")
        h, m = int(parts[0]), int(parts[1])
        return datetime(1900, 1, 1, h % 24, m).time()
    except Exception:
        return None

def _min_to_time(minutes):
    """Float minutes → datetime.time"""
    if minutes is None:
        return None
    try:
        total = int(float(minutes))
        h, m = divmod(total, 60)
        return datetime(1900, 1, 1, h % 24, m).time()
    except Exception:
        return None

# ── Main export function ──────────────────────────────────────
def build_excel(saisies_with_tours: list) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suivi Irrigation"

    # ── Column widths (exact from template) ──────────────────
    ws.column_dimensions["A"].width = 20.0
    ws.column_dimensions["B"].width = 9.66
    ws.column_dimensions["C"].width = 9.0
    ws.column_dimensions["D"].width = 13.0
    ws.column_dimensions["E"].width = 9.22
    ws.column_dimensions["F"].width = 25.11
    for i in range(7, 27):   # G to Z
        ws.column_dimensions[get_column_letter(i)].width = 7.5
    ws.column_dimensions["AA"].width = 26.44
    ws.column_dimensions["AB"].width = 12.11

    # ── Header row ───────────────────────────────────────────
    header_fill   = _fill(DARK_HEADER)
    header_font   = _font(bold=True, size=11, color="FFFFFFFF")
    header_border = _border_thin()

    for col, label in enumerate(["Date", "Ferme", "Station", "Serre", "Vanne"], start=1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = header_font; c.fill = header_fill
        c.alignment = _align(); c.border = header_border

    # Tours label + numbers 1-20
    c = ws.cell(row=1, column=6, value="Tours")
    c.font = _font(bold=True, size=10, color="FFFFFFFF")
    c.fill = header_fill; c.alignment = _align(); c.border = _border_label()

    for t in range(1, 21):
        c = ws.cell(row=1, column=6 + t, value=t)
        c.font = _font(bold=True, size=10, color="FFFFFFFF")
        c.fill = header_fill; c.alignment = _align(); c.border = header_border

    # Empty AA/AB header
    for col in [27, 28]:
        c = ws.cell(row=1, column=col, value="")
        c.fill = header_fill; c.border = header_border

    ws.row_dimensions[1].height = 18

    # ── Row layout definition ─────────────────────────────────
    # (idx, label, label_fill, data_fill, aa_label, is_time, is_pct, saisie_only)
    # saisie_only=True → valeur écrite uniquement en col G (pas répétée sur les tours)
    ROW_DEF = [
        (0,  "Radiation",          GREEN_LIGHT, GREEN_LIGHT, "Nombre De Bras",       False, False, False),
        (1,  "Cumul Radiation",    GREEN_LIGHT, GREEN_LIGHT, "Nombre de Goutteurs",  False, False, False),
        (2,  "Heure",              GREEN_MED,   None,        "Durée totale",         True,  False, False),
        (3,  "Durée (min)",        GREEN_MED,   None,        "EC Bassin",            True,  False, False),
        (4,  "Temps Repos (min)",  GREEN_MED,   None,        "Total V Apport",       True,  False, False),
        (5,  "V Apport (cc)",      GREEN_MED,   None,        "EC cumul apport",      False, False, False),
        (6,  "EC Apport",          GREEN_MED,   None,        "PH cumul apport",      False, False, False),
        (7,  "pH Apport",          GREEN_MED,   None,        "Total V Drainage",     False, False, False),
        (8,  "V Drainage (cc)",    GREEN_MED,   None,        "Moyenne % drainage",   False, False, False),
        (9,  "% Drainage",         GREEN_MED,   None,        "EC cumul drainage",    False, True,  False),
        (10, "Moyenne % Drainage", GREEN_MED,   None,        "PH cumul Drainage",    False, True,  False),
        (11, "EC Drainage",        GREEN_MED,   None,        "CC/bras consommé",     False, False, False),
        (12, "pH Drainage",        GREEN_MED,   None,        "Nombre des tours",     False, False, False),
        (13, "Heure Matin",        ORANGE,      ORANGE,      None,                   True,  False, True),
        (14, "Heure Soir",         ORANGE,      ORANGE,      None,                   True,  False, True),
        (15, "% Ressuyage",        ORANGE,      ORANGE,      None,                   False, True,  True),
        (16, "Poids Matin (Kg)",   ORANGE,      ORANGE,      None,                   False, False, True),
        (17, "Poids Soir (Kg)",    ORANGE,      ORANGE,      None,                   False, False, True),
        (18, "EC Bassin",          ORANGE,      ORANGE,      None,                   False, False, True),
    ]

    current_row = 2

    for saisie, tours in saisies_with_tours:
        # Sort tours by num_tour
        tours = sorted(tours, key=lambda t: t["num_tour"])
        max_tours = min(len(tours), 20)

        # Precompute summary values
        sum_map = {
            "Nombre De Bras":      saisie["nbr_bras"],
            "Nombre de Goutteurs": saisie["nbr_goutteurs"],
            "Durée totale":        _parse_time_str(saisie.get("duree_totale")),
            "EC Bassin":           saisie["bassin_ec"],
            "Total V Apport":      saisie["total_v_apport"],
            "EC cumul apport":     saisie["ec_moy_apport"],
            "PH cumul apport":     saisie["ph_moy_apport"],
            "Total V Drainage":    saisie["total_v_drain"],
            "Moyenne % drainage":  (saisie["moy_drain_finale"] / 100.0
                                    if saisie.get("moy_drain_finale") is not None else None),
            "EC cumul drainage":   saisie["ec_moy_drain"],
            "PH cumul Drainage":   saisie["ph_moy_drain"],
            "CC/bras consommé":    saisie["cc_bras"],
            "Nombre des tours":    saisie["nbr_tours"],
        }

        # Data per row index
        def tour_value(row_idx, t):
            m = {
                0:  t.get("rad"),
                1:  t.get("cumul_rad"),
                2:  _parse_time_str(t.get("heure")),
                3:  _min_to_time(t.get("duree_min")),
                4:  _min_to_time(t.get("temps_repos")),
                5:  t.get("v_apport"),
                6:  t.get("ec_apport"),
                7:  t.get("ph_apport"),
                8:  t.get("v_drain"),
                9:  (t["pct_drain"] / 100.0 if t.get("pct_drain") is not None else None),
                10: (t["moy_pct_drain"] / 100.0 if t.get("moy_pct_drain") is not None else None),
                11: t.get("ec_drain"),
                12: t.get("ph_drain"),
                13: _parse_time_str(saisie.get("heure_matin")),
                14: _parse_time_str(saisie.get("heure_soir")),
                15: (saisie["pct_ressuyage"] / 100.0 if saisie.get("pct_ressuyage") is not None else None),
                16: saisie.get("poids_matin"),
                17: saisie.get("poids_soir"),
                18: saisie.get("bassin_ec"),
            }
            return m.get(row_idx)

        for rd in ROW_DEF:
            idx, label, label_fill, data_fill, aa_label, is_time, is_pct, saisie_only = rd
            r = current_row + idx

            ws.row_dimensions[r].height = 15

            # A-E: identity cells
            date_val = saisie["date"]
            if isinstance(date_val, str):
                try:
                    date_val = datetime.strptime(date_val, "%Y-%m-%d").date()
                except Exception:
                    pass

            for col, val in enumerate([date_val, saisie["farm_name"], saisie["station"],
                                        saisie["serre"], saisie["vanne"]], start=1):
                sz = 11 if col <= 5 else 10
                _set(ws, r, col, val, bold=True, size=sz)
                if col == 1 and isinstance(val, (date, datetime)):
                    ws.cell(r, col).number_format = "dd/mm/yyyy"

            # F: label
            _set(ws, r, 6, label, fill_hex=label_fill, bold=True, size=10,
                 border=_border_label())

            # G-Z: tour data
            if saisie_only:
                # Écrire uniquement en col G (valeur de la saisie, pas par tour)
                val = tour_value(idx, tours[0] if tours else {})
                cell = _set(ws, r, 7, val, fill_hex=data_fill, bold=True, size=10,
                            border=_border_thin() if data_fill else None)
                if val is not None:
                    if is_time:
                        ws.cell(r, 7).number_format = "hh:mm"
                    elif is_pct:
                        ws.cell(r, 7).number_format = "0%"
                    else:
                        ws.cell(r, 7).number_format = "0.00"
                # Cols H-Z vides mais avec fill si data_fill
                for ti in range(1, 20):
                    col = 7 + ti
                    _set(ws, r, col, None, fill_hex=None, bold=True, size=10)
            else:
                for ti in range(20):
                    col = 7 + ti
                    val = None
                    if ti < len(tours):
                        val = tour_value(idx, tours[ti])
                    cell = _set(ws, r, col, val, fill_hex=data_fill, bold=True, size=10,
                                border=_border_thin() if data_fill else None)
                    if val is not None:
                        if is_time:
                            ws.cell(r, col).number_format = "hh:mm"
                        elif is_pct:
                            ws.cell(r, col).number_format = "0%"
                        else:
                            ws.cell(r, col).number_format = "0.00"

            # AA / AB: summary
            if aa_label:
                # AA label
                _set(ws, r, 27, aa_label, fill_hex=ORANGE, bold=True, size=11,
                     border=_border_thin())
                # AB value
                ab_val = sum_map.get(aa_label)
                ab_fmt = None
                if aa_label == "Durée totale" and isinstance(ab_val, type(datetime.now().time())):
                    ab_fmt = "hh:mm"
                elif aa_label == "Moyenne % drainage" and ab_val is not None:
                    ab_fmt = "0%"
                c = _set(ws, r, 28, ab_val, fill_hex=None, bold=True, size=11,
                         border=_border_thin())
                if ab_fmt:
                    c.number_format = ab_fmt
            else:
                # Empty AA/AB
                for col in [27, 28]:
                    ws.cell(r, col).border = _border_thin()

        current_row += 19

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Endpoint ──────────────────────────────────────────────────
@router.get("/saisie")
async def export_saisie_excel(
    farm_names: str = Query(..., description="Fermes séparées par virgule"),
    date_from:  str = Query(...),
    date_to:    str = Query(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    import re

    farms_raw = [f.strip() for f in farm_names.split(",") if f.strip()]
    if not farms_raw:
        raise HTTPException(400, "Aucune ferme spécifiée")
    
    # Validation stricte du format farm_name
    FARM_NAME_RE = re.compile(r'^[a-zA-Z0-9_\- ]{1,50}$')
    farms = [f for f in farms_raw if FARM_NAME_RE.match(f)]
    if not farms:
        raise HTTPException(400, "Format de nom de ferme invalide")
    if len(farms) > 20:
        raise HTTPException(400, "Trop de fermes demandées (max 20)")

    # Sécurité : opérateur ne peut exporter que ses fermes
    if current_user["role"] != "admin":
        allowed = current_user.get("farm_names") or []
        farms = [f for f in farms if f in allowed]
        if not farms:
            raise HTTPException(403, "Accès refusé aux fermes demandées")
    
    try:
        date_from_d = datetime.strptime(date_from, "%Y-%m-%d").date()
        date_to_d   = datetime.strptime(date_to,   "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, "Format de date invalide (YYYY-MM-DD)")
    
    # Limiter la plage à 1 an
    from datetime import timedelta
    if (date_to_d - date_from_d).days > 366:
        raise HTTPException(400, "Plage de dates limitée à 1 an")
    if date_from_d > date_to_d:
        raise HTTPException(400, "date_from doit être avant date_to")

    saisies_query = (
        db.query(SaisieJournaliere)
        .filter(
            SaisieJournaliere.farm_name.in_(farms),
            SaisieJournaliere.date.between(date_from_d, date_to_d)
        )
        .order_by(
            SaisieJournaliere.farm_name,
            SaisieJournaliere.date,
            SaisieJournaliere.station,
            SaisieJournaliere.serre,
            SaisieJournaliere.vanne,
        )
        .all()
    )
    saisies = [dict(s.__dict__) for s in saisies_query]

    if not saisies:
        raise HTTPException(404, "Aucune donnée pour les critères sélectionnés")

    # ── Fetch tours ──────────────────────────────────────────
    saisie_ids = [s["id"] for s in saisies]
    result2 = db.execute(text("""
        SELECT * FROM saisie_tours
        WHERE saisie_id = ANY(:ids)
        ORDER BY saisie_id, num_tour
    """), {"ids": saisie_ids})
    tours_raw = [dict(r._mapping) for r in result2]

    tours_by_saisie = {}
    for t in tours_raw:
        tours_by_saisie.setdefault(t["saisie_id"], []).append(t)

    saisies_with_tours = [
        (s, tours_by_saisie.get(s["id"], []))
        for s in saisies
    ]

    # ── Build Excel dans un thread (openpyxl bloque l'event loop) ──
    import asyncio
    from functools import partial
    from loguru import logger

    loop = asyncio.get_event_loop()
    try:
        excel_buffer = await loop.run_in_executor(
            None, partial(build_excel, saisies_with_tours)
        )
    except Exception as e:
        logger.error(f"Erreur build_excel : {e}")
        raise HTTPException(500, f"Erreur génération Excel : {str(e)}")

    filename = f"suivi_irrigation_{date_from}_{date_to}.xlsx"

    # Lire tout le contenu en bytes pour éviter les problèmes de streaming
    excel_bytes = excel_buffer.read()
    if not excel_bytes:
        raise HTTPException(500, "Fichier Excel vide généré")

    from fastapi.responses import Response
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(excel_bytes)),
        },
    )